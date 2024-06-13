#pip install cx_Oracle
#install oracle instant client.



import cx_Oracle
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

def send_mail():
    try:
        date = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")
        conn_str = 'PPL1/ppl@178.128.217.191:1521/MALAYSIA'

        conn = cx_Oracle.connect(conn_str)
        cursor = conn.cursor()

        query = f"SELECT COUNTRY, ACMP_NAME, ACCH_NAME AS ACCOUNT_HEAD_NAME, JDT, DESD AS DESCRIPTION, PAYMENT AS EXPENSE FROM PPL1.TT WHERE JDT = '{date}'"
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [i[0] for i in cursor.description]
        df = pd.DataFrame(rows, columns=columns)

        total_expense = df['EXPENSE'].sum()

        attachment_path = get_attachment(df)

        email = 'automation@mis.prangroup.com'
        epass = 'aaaaAAAA0000'

        msg = MIMEMultipart()
        msg['From'] = email
        msg['To'] = 'mis33@mis.prangroup.com'
        msg['Subject'] = f'LAST DAY EXPENSE | IBS | {date}'

        body = """\
            <html>
              <body>
                <div style='font-family:calibri;font-size:15px;'>
                  This is a computer-generated mail. Please do not reply.
                </div>
                {}
              </body>
            </html>
            """.format(df.to_html(index=False))

        msg.attach(MIMEText(body, 'html'))

        attachment = open(attachment_path, 'rb')
        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= {}".format(attachment_path.split('/')[-1]))
        msg.attach(part)

        smtp_server = 'mail.mis.prangroup.com'
        server = smtplib.SMTP(smtp_server, 25)
        server.starttls()
        server.login(email, epass)
        server.sendmail(email, 'mis33@mis.prangroup.com', msg.as_string())
        server.quit()

    except Exception as e:
        print(str(e))

def get_attachment(dataframe):
    file_path = f'IBS_FINANCIAL_REPORT_{(datetime.now() - timedelta(days=1)).strftime("%d-%m-%Y")}.xlsx'

    wb = Workbook()
    ws = wb.active

    # Writing DataFrame to Excel
    for r in dataframe.to_numpy():
        ws.append(list(r))

    # Applying styles
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            cell.font = Font(bold=True)
            #cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    wb.save(file_path)

    return file_path

if __name__ == "__main__":
    send_mail()
